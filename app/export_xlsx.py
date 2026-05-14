"""Export all pipeline data from Postgres to an Excel workbook.

Produces the same tabs as the Google Sheet (Output, Match Results,
Active Threads, Products Pipeline, Dashboard, Automation Stats,
Thread Activity) but writes to a local .xlsx file.

Usage:
    python -m app.export_xlsx [output_path]
    docker exec <container> python -m app.export_xlsx /tmp/export.xlsx

Default output: exports/sourcing_export_YYYY-MM-DD.xlsx
"""

import sys
from collections import Counter
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func as sa_func

from app.db.database import SessionLocal
from app.db.models.automation_event import AutomationEvent
from app.db.models.quote import Quote
from app.db.models.source_product import SourceProduct
from app.db.models.supplier_product import SupplierProduct
from app.db.models.supplier_thread import VALID_STATES, SupplierThread
from app.services.sheets import (
    ACTIVE_THREADS_HEADERS,
    AUTOMATION_STATS_HEADERS,
    DASHBOARD_HEADERS,
    MATCH_RESULTS_HEADERS,
    OUTPUT_COLUMNS,
    PRODUCTS_PIPELINE_HEADERS,
    THREAD_ACTIVITY_HEADERS,
)

GMAIL_THREAD_URL = "https://mail.google.com/mail/u/0/#inbox/{}"
THREAD_ACTIVITY_LIMIT = 1000
AUTOMATION_HEALTH_DAYS = 7
TOP_ERRORS_LIMIT = 10

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _style_headers(ws):
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _add_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list]) -> None:
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _style_headers(ws)


def build_output_rows(session) -> list[list]:
    threads = (
        session.query(SupplierThread)
        .filter(SupplierThread.state != "NEW")
        .order_by(SupplierThread.created_at.desc())
        .all()
    )
    rows = []
    for t in threads:
        source = t.source_product
        supplier = t.supplier
        latest_quote = t.quotes[-1] if t.quotes else None
        first_outbound = next(
            (m for m in t.messages if m.direction == "outbound"), None
        )
        gmail_link = (
            GMAIL_THREAD_URL.format(t.gmail_thread_id) if t.gmail_thread_id else ""
        )

        rows.append(
            [
                source.title,
                source.url,
                source.slug,
                supplier.name,
                f"{latest_quote.price_usd:.2f}" if latest_quote else "Awaiting Quotes",
                str(latest_quote.moq) if latest_quote and latest_quote.moq else "",
                latest_quote.lead_time or "" if latest_quote else "",
                gmail_link,
                t.last_updated.strftime("%Y-%m-%d") if t.last_updated else "",
                (
                    first_outbound.sent_at.strftime("%Y-%m-%d")
                    if first_outbound and first_outbound.sent_at
                    else ""
                ),
                t.supplier_product.platform,
            ]
        )
    return rows


def build_match_results_rows(session) -> list[list]:
    products = (
        session.query(SupplierProduct)
        .order_by(
            SupplierProduct.source_product_id,
            SupplierProduct.match_status.desc(),
            SupplierProduct.match_confidence.desc(),
        )
        .all()
    )
    return [
        [
            sp.source_product.title,
            sp.title,
            sp.supplier.name,
            sp.platform,
            sp.match_status,
            f"{sp.match_confidence:.2f}" if sp.match_confidence is not None else "",
            sp.match_reason or "",
            sp.product_url,
        ]
        for sp in products
    ]


def build_active_threads_rows(session) -> list[list]:
    now = datetime.now(timezone.utc)
    threads = (
        session.query(SupplierThread)
        .filter(SupplierThread.state.notin_(("CLOSED", "UNPROCESSABLE")))
        .order_by(SupplierThread.last_updated.desc())
        .all()
    )
    rows = []
    for t in threads:
        first_outbound = next(
            (m for m in t.messages if m.direction == "outbound"), None
        )
        days_since = ""
        if first_outbound and first_outbound.sent_at:
            days_since = str((now - first_outbound.sent_at).days)

        msgs_sent = sum(1 for m in t.messages if m.direction == "outbound")
        msgs_received = sum(1 for m in t.messages if m.direction == "inbound")
        quotes_received = len(t.quotes)
        latest_quote = f"{t.quotes[-1].price_usd:.2f}" if t.quotes else ""
        best_quote = f"{min(q.price_usd for q in t.quotes):.2f}" if t.quotes else ""

        link = ""
        if t.gmail_thread_id:
            link = GMAIL_THREAD_URL.format(t.gmail_thread_id)
        elif t.platform_thread_url:
            link = t.platform_thread_url

        respond_after = ""
        if t.respond_after:
            respond_after = t.respond_after.strftime("%Y-%m-%d %H:%M:%S")

        rows.append(
            [
                str(t.id),
                t.source_product.title,
                t.source_product.url,
                t.supplier.name,
                t.supplier.platform,
                t.channel,
                t.state,
                days_since,
                str(msgs_sent),
                str(msgs_received),
                str(quotes_received),
                latest_quote,
                best_quote,
                str(t.negotiation_rounds),
                respond_after,
                link,
            ]
        )
    return rows


def build_automation_stats_rows(session) -> list[list]:
    results = (
        session.query(
            AutomationEvent.stage,
            AutomationEvent.action,
            AutomationEvent.outcome,
            sa_func.count().label("count"),
            sa_func.max(AutomationEvent.created_at).label("latest"),
        )
        .group_by(
            AutomationEvent.stage,
            AutomationEvent.action,
            AutomationEvent.outcome,
        )
        .order_by(
            AutomationEvent.stage,
            AutomationEvent.action,
            AutomationEvent.outcome,
        )
        .all()
    )
    return [
        [
            stage,
            action,
            outcome,
            str(count),
            latest.strftime("%Y-%m-%d %H:%M:%S") if latest else "",
        ]
        for stage, action, outcome, count, latest in results
    ]


def build_dashboard_rows(session) -> list[list]:
    rows = []
    now = datetime.now(timezone.utc)
    cutoff = now - timedelta(days=AUTOMATION_HEALTH_DAYS)

    total_products = session.query(sa_func.count(SourceProduct.id)).scalar()
    products_with_match = (
        session.query(
            sa_func.count(sa_func.distinct(SupplierProduct.source_product_id))
        )
        .filter(SupplierProduct.match_status == "matched")
        .scalar()
    )
    products_with_active = (
        session.query(sa_func.count(sa_func.distinct(SupplierThread.source_product_id)))
        .filter(SupplierThread.state.notin_(("CLOSED", "UNPROCESSABLE")))
        .scalar()
    )
    total_contacted = (
        session.query(sa_func.count(SupplierThread.id))
        .filter(SupplierThread.state != "NEW")
        .scalar()
    )
    threads_with_quote = session.query(
        sa_func.count(sa_func.distinct(Quote.thread_id))
    ).scalar()

    rows.append(["Pipeline Summary", "Total Products", str(total_products), ""])
    rows.append(
        [
            "Pipeline Summary",
            "Products with Matched Supplier",
            str(products_with_match),
            "",
        ]
    )
    rows.append(
        [
            "Pipeline Summary",
            "Products with Active Thread",
            str(products_with_active),
            "",
        ]
    )
    rows.append(
        ["Pipeline Summary", "Total Suppliers Contacted", str(total_contacted), ""]
    )
    rows.append(
        ["Pipeline Summary", "Threads with Quote Received", str(threads_with_quote), ""]
    )

    searched_per_product = (
        session.query(
            SupplierProduct.source_product_id,
            sa_func.count(SupplierProduct.id).label("cnt"),
        )
        .group_by(SupplierProduct.source_product_id)
        .subquery()
    )
    avg_searched = session.query(sa_func.avg(searched_per_product.c.cnt)).scalar()

    matched_per_product = (
        session.query(
            SupplierProduct.source_product_id,
            sa_func.count(SupplierProduct.id).label("cnt"),
        )
        .filter(SupplierProduct.match_status == "matched")
        .group_by(SupplierProduct.source_product_id)
        .subquery()
    )
    avg_matched = session.query(sa_func.avg(matched_per_product.c.cnt)).scalar()
    rows.append(
        [
            "Pipeline Summary",
            "Avg Supplier Products Searched",
            f"{avg_searched:.1f}" if avg_searched else "0",
            "",
        ]
    )
    rows.append(
        [
            "Pipeline Summary",
            "Avg Supplier Products Matched",
            f"{avg_matched:.1f}" if avg_matched else "0",
            "",
        ]
    )

    state_counts = dict(
        session.query(SupplierThread.state, sa_func.count(SupplierThread.id))
        .group_by(SupplierThread.state)
        .all()
    )
    for state in VALID_STATES:
        rows.append(["Thread Funnel", state, str(state_counts.get(state, 0)), ""])

    health_results = (
        session.query(
            AutomationEvent.stage,
            AutomationEvent.action,
            AutomationEvent.outcome,
            sa_func.count().label("count"),
        )
        .filter(AutomationEvent.created_at >= cutoff)
        .group_by(
            AutomationEvent.stage,
            AutomationEvent.action,
            AutomationEvent.outcome,
        )
        .all()
    )
    health_by_pair: dict[tuple[str, str], Counter] = {}
    for stage, action, outcome, count in health_results:
        key = (stage, action)
        if key not in health_by_pair:
            health_by_pair[key] = Counter()
        health_by_pair[key][outcome] = count

    for (stage, action), counts in sorted(health_by_pair.items()):
        total = sum(counts.values())
        parts = []
        for outcome in ("deterministic", "agent_fallback", "failed"):
            c = counts.get(outcome, 0)
            pct = c / total * 100 if total else 0
            parts.append(f"{outcome}: {c} ({pct:.0f}%)")
        rows.append(
            [
                f"Automation Health ({AUTOMATION_HEALTH_DAYS}d)",
                f"{stage} / {action}",
                "  ".join(parts),
                str(total),
            ]
        )

    total_failed_7d = (
        session.query(sa_func.count(AutomationEvent.id))
        .filter(
            AutomationEvent.created_at >= cutoff,
            AutomationEvent.outcome == "failed",
        )
        .scalar()
    ) or 0

    error_results = (
        session.query(
            AutomationEvent.detail,
            AutomationEvent.stage,
            AutomationEvent.action,
            sa_func.count().label("count"),
        )
        .filter(
            AutomationEvent.created_at >= cutoff,
            AutomationEvent.outcome == "failed",
            AutomationEvent.detail.isnot(None),
        )
        .group_by(
            AutomationEvent.detail,
            AutomationEvent.stage,
            AutomationEvent.action,
        )
        .order_by(sa_func.count().desc())
        .limit(TOP_ERRORS_LIMIT)
        .all()
    )
    for detail, stage, action, count in error_results:
        pct = count / total_failed_7d * 100 if total_failed_7d else 0
        rows.append(
            [
                f"Top Errors ({AUTOMATION_HEALTH_DAYS}d)",
                f"{stage} / {action}",
                f"{count} ({pct:.1f}% of failures)",
                detail,
            ]
        )

    avg_rounds = (
        session.query(sa_func.avg(SupplierThread.negotiation_rounds))
        .filter(SupplierThread.negotiation_rounds > 0)
        .scalar()
    )
    total_quotes = session.query(sa_func.count(Quote.id)).scalar()
    rows.append(
        [
            "Quote Stats",
            "Avg Negotiation Rounds",
            f"{avg_rounds:.1f}" if avg_rounds else "0",
            "",
        ]
    )
    rows.append(["Quote Stats", "Total Quotes Received", str(total_quotes), ""])
    rows.append(["Last Sync", "Timestamp", now.strftime("%Y-%m-%d %H:%M:%S UTC"), ""])

    return rows


def build_products_pipeline_rows(session) -> list[list]:
    products = (
        session.query(SourceProduct).order_by(SourceProduct.created_at.desc()).all()
    )
    rows = []
    for p in products:
        candidates = (
            session.query(SupplierProduct)
            .filter(SupplierProduct.source_product_id == p.id)
            .all()
        )
        status_counts = Counter(c.match_status for c in candidates)

        threads = (
            session.query(SupplierThread)
            .filter(SupplierThread.source_product_id == p.id)
            .all()
        )
        active_threads = [
            t for t in threads if t.state not in ("CLOSED", "UNPROCESSABLE")
        ]
        threads_with_quote = [t for t in threads if t.quotes]

        best_quote = ""
        all_prices = [q.price_usd for t in threads for q in t.quotes]
        if all_prices:
            best_quote = f"{min(all_prices):.2f}"

        rows.append(
            [
                p.title,
                p.url,
                str(len(candidates)),
                str(status_counts.get("matched", 0)),
                str(status_counts.get("rejected", 0)),
                str(status_counts.get("pending", 0)),
                str(len(active_threads)),
                str(len(threads_with_quote)),
                best_quote,
                "",
            ]
        )
    return rows


def build_thread_activity_rows(session) -> list[list]:
    events = (
        session.query(AutomationEvent)
        .order_by(AutomationEvent.created_at.desc())
        .limit(THREAD_ACTIVITY_LIMIT)
        .all()
    )

    thread_ids = {e.supplier_thread_id for e in events if e.supplier_thread_id}
    threads_by_id: dict[int, SupplierThread] = {}
    if thread_ids:
        thread_rows = (
            session.query(SupplierThread)
            .filter(SupplierThread.id.in_(thread_ids))
            .all()
        )
        threads_by_id = {t.id: t for t in thread_rows}

    rows = []
    for e in events:
        thread = (
            threads_by_id.get(e.supplier_thread_id) if e.supplier_thread_id else None
        )
        rows.append(
            [
                e.created_at.strftime("%Y-%m-%d %H:%M:%S") if e.created_at else "",
                e.stage,
                e.action,
                e.outcome,
                str(e.supplier_thread_id) if e.supplier_thread_id else "",
                thread.source_product.title if thread else "",
                thread.supplier.name if thread else "",
                thread.channel if thread else "",
                e.detail or "",
            ]
        )
    return rows


def main():
    output_dir = Path("exports")
    output_dir.mkdir(exist_ok=True)

    if len(sys.argv) > 1:
        output_path = Path(sys.argv[1])
    else:
        date_str = datetime.now().strftime("%Y-%m-%d")
        output_path = output_dir / f"sourcing_export_{date_str}.xlsx"

    print(f"Exporting pipeline data to {output_path} ...")

    wb = Workbook()
    wb.remove(wb.active)

    with SessionLocal() as session:
        output_headers = list(OUTPUT_COLUMNS.keys())
        output_headers_display = [h.replace("_", " ").title() for h in output_headers]

        print("  Building Output tab ...")
        output_rows = build_output_rows(session)
        _add_sheet(wb, "Output", output_headers_display, output_rows)
        print(f"    {len(output_rows)} rows")

        print("  Building Match Results tab ...")
        match_rows = build_match_results_rows(session)
        _add_sheet(wb, "Match Results", MATCH_RESULTS_HEADERS, match_rows)
        print(f"    {len(match_rows)} rows")

        print("  Building Active Threads tab ...")
        active_rows = build_active_threads_rows(session)
        _add_sheet(wb, "Active Threads", ACTIVE_THREADS_HEADERS, active_rows)
        print(f"    {len(active_rows)} rows")

        print("  Building Automation Stats tab ...")
        stats_rows = build_automation_stats_rows(session)
        _add_sheet(wb, "Automation Stats", AUTOMATION_STATS_HEADERS, stats_rows)
        print(f"    {len(stats_rows)} rows")

        print("  Building Dashboard tab ...")
        dashboard_rows = build_dashboard_rows(session)
        _add_sheet(wb, "Dashboard", DASHBOARD_HEADERS, dashboard_rows)
        print(f"    {len(dashboard_rows)} rows")

        print("  Building Products Pipeline tab ...")
        pipeline_rows = build_products_pipeline_rows(session)
        _add_sheet(wb, "Products Pipeline", PRODUCTS_PIPELINE_HEADERS, pipeline_rows)
        print(f"    {len(pipeline_rows)} rows")

        print("  Building Thread Activity tab ...")
        activity_rows = build_thread_activity_rows(session)
        _add_sheet(wb, "Thread Activity", THREAD_ACTIVITY_HEADERS, activity_rows)
        print(f"    {len(activity_rows)} rows")

    wb.save(output_path)
    print(f"\nDone! Exported to {output_path}")


if __name__ == "__main__":
    main()
